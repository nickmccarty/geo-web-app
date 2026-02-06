"""
Utility functions for TIF processing and GeoJSON handling.
"""
import os
os.environ['KMP_DUPLICATE_LIB_OK'] = 'TRUE'
import io
import json
import random
from typing import Tuple
import numpy as np
import pandas as pd
from PIL import Image
import rasterio
from rasterio.warp import calculate_default_transform, reproject, Resampling
from rasterio.windows import from_bounds
import geopandas as gpd
from shapely.ops import unary_union


def generate_preview_image(tif_path: str, max_size: int = 1024) -> Tuple[bytes, dict]:
    """
    Generate a downsampled PNG preview of GeoTIFF for web display.

    Args:
        tif_path: Path to GeoTIFF file
        max_size: Maximum dimension for preview

    Returns:
        Tuple of (PNG bytes, metadata dict)
    """
    with rasterio.open(tif_path) as src:
        # Read RGB bands
        data = src.read([1, 2, 3])

        # Get metadata
        metadata = {
            'width': src.width,
            'height': src.height,
            'crs': str(src.crs),
            'bounds': list(src.bounds)
        }

        # Calculate resize factor
        h, w = data.shape[1], data.shape[2]
        scale = max_size / max(h, w)

        if scale < 1:
            new_h = int(h * scale)
            new_w = int(w * scale)
        else:
            new_h, new_w = h, w

        # Convert to PIL Image and resize
        img_array = np.moveaxis(data, 0, 2)  # CHW -> HWC
        img = Image.fromarray(img_array.astype(np.uint8))
        img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)

        # Convert to PNG bytes
        img_buffer = io.BytesIO()
        img.save(img_buffer, format='PNG')
        img_bytes = img_buffer.getvalue()

    return img_bytes, metadata


def generate_overlay_image(tif_path: str, max_size: int = 8192, format: str = 'JPEG', quality: int = 85) -> bytes:
    """
    Generate a high-resolution overlay image from GeoTIFF for map display.
    Reprojects to WGS84 (EPSG:4326) to match Leaflet's coordinate system.

    Args:
        tif_path: Path to GeoTIFF file
        max_size: Maximum dimension for overlay (default 8192px)
        format: Output format ('JPEG' or 'PNG')
        quality: JPEG quality (1-100, only used for JPEG)

    Returns:
        Image bytes in specified format
    """
    with rasterio.open(tif_path) as src:
        # Calculate transform to WGS84
        dst_crs = 'EPSG:4326'
        transform, width, height = calculate_default_transform(
            src.crs, dst_crs, src.width, src.height, *src.bounds
        )

        # Apply max_size constraint
        scale = max_size / max(height, width)
        if scale < 1:
            width = int(width * scale)
            height = int(height * scale)
            # Adjust transform for new dimensions
            transform = transform * transform.scale(1/scale, 1/scale)

        # Create destination array for reprojected data
        dst_data = np.zeros((3, height, width), dtype=np.uint8)

        # Reproject each RGB band
        for i in range(3):
            reproject(
                source=rasterio.band(src, i + 1),
                destination=dst_data[i],
                src_transform=src.transform,
                src_crs=src.crs,
                dst_transform=transform,
                dst_crs=dst_crs,
                resampling=Resampling.lanczos
            )

        # Convert to PIL Image
        img_array = np.moveaxis(dst_data, 0, 2)  # CHW -> HWC
        img = Image.fromarray(img_array.astype(np.uint8))

        # Convert to bytes
        img_buffer = io.BytesIO()
        if format.upper() == 'JPEG':
            img.save(img_buffer, format='JPEG', quality=quality, optimize=True)
        else:
            img.save(img_buffer, format='PNG', optimize=True)

        img_bytes = img_buffer.getvalue()

    return img_bytes


def get_tif_bounds_geojson(tif_path: str) -> dict:
    """
    Get GeoTIFF bounds as GeoJSON for map initialization.

    Args:
        tif_path: Path to GeoTIFF file

    Returns:
        GeoJSON feature collection with bounds polygon
    """
    with rasterio.open(tif_path) as src:
        bounds = src.bounds
        crs = src.crs

        # Create bounds polygon
        from shapely.geometry import box
        bounds_poly = box(bounds.left, bounds.bottom, bounds.right, bounds.top)

        # Create GeoDataFrame and convert to WGS84 for Leaflet
        gdf = gpd.GeoDataFrame([{'geometry': bounds_poly}], crs=crs)
        gdf_wgs84 = gdf.to_crs('EPSG:4326')

        # Convert to GeoJSON
        return json.loads(gdf_wgs84.to_json())


def merge_overlapping_geometries(gdf: gpd.GeoDataFrame, buffer_dist: float = 0.1) -> gpd.GeoDataFrame:
    """
    Merge overlapping or nearby geometries in GeoDataFrame.

    Args:
        gdf: GeoDataFrame with detections
        buffer_dist: Distance for buffering/merging (in CRS units)

    Returns:
        GeoDataFrame with merged geometries
    """
    if gdf is None or len(gdf) == 0:
        return gdf

    # Buffer slightly to merge nearby boxes
    buffered = gdf.geometry.buffer(buffer_dist)

    # Union all overlapping geometries
    merged = unary_union(buffered)

    # Unbuffer to get original size back
    if hasattr(merged, 'geoms'):
        # MultiPolygon
        final_geoms = [geom.buffer(-buffer_dist) for geom in merged.geoms]
    else:
        # Single Polygon
        final_geoms = [merged.buffer(-buffer_dist)]

    # Create new GeoDataFrame
    merged_gdf = gpd.GeoDataFrame(
        {'geometry': final_geoms, 'class': 'object'},
        crs=gdf.crs
    )

    return merged_gdf


def gdf_to_geojson_for_leaflet(gdf: gpd.GeoDataFrame) -> dict:
    """
    Convert GeoDataFrame to GeoJSON in WGS84 for Leaflet display.

    Args:
        gdf: GeoDataFrame in any CRS

    Returns:
        GeoJSON dict in EPSG:4326
    """
    if gdf is None or len(gdf) == 0:
        return {"type": "FeatureCollection", "features": []}

    # Convert to WGS84
    gdf_wgs84 = gdf.to_crs('EPSG:4326')

    # Convert to GeoJSON
    return json.loads(gdf_wgs84.to_json())


def load_qc_points(csv_bytes: bytes) -> pd.DataFrame:
    """
    Parse QC check points CSV. Supports both header-based and positional column mapping.

    Header-based: recognizes columns named y, x, point_id, z (any order).
    Positional fallback: assumes column order point_id, y, x, z.

    Args:
        csv_bytes: Raw CSV file bytes

    Returns:
        DataFrame with columns: point_id, y, x, z
    """
    # First pass: check if first row is a header
    raw = pd.read_csv(io.BytesIO(csv_bytes), header=None, dtype=str)

    if len(raw.columns) < 2:
        raise ValueError(f"CSV must have at least 2 columns (y, x), found {len(raw.columns)}")

    # Check if first row contains recognizable header names
    first_row_lower = [str(v).strip().lower() for v in raw.iloc[0]]
    known_headers = {'y', 'x', 'z', 'point_id', 'northing', 'easting', 'elev', 'id'}
    has_header = any(val in known_headers for val in first_row_lower)

    if has_header:
        # Re-read with header
        df = pd.read_csv(io.BytesIO(csv_bytes), dtype=str)
        df.columns = [c.strip().lower() for c in df.columns]

        # Map common aliases
        col_map = {}
        for c in df.columns:
            if c in ('y', 'northing'):
                col_map['y'] = c
            elif c in ('x', 'easting'):
                col_map['x'] = c
            elif c in ('point_id', 'id'):
                col_map['point_id'] = c
            elif c in ('z', 'elev', 'elevation'):
                col_map['z'] = c

        if 'y' not in col_map or 'x' not in col_map:
            raise ValueError(
                f"CSV header must include 'y' and 'x' columns. Found: {list(df.columns)}"
            )

        out = pd.DataFrame({
            'y': pd.to_numeric(df[col_map['y']], errors='coerce'),
            'x': pd.to_numeric(df[col_map['x']], errors='coerce'),
            'point_id': pd.to_numeric(
                df[col_map['point_id']].astype(str).str.replace(r'^\s*CK\s*', '', regex=True).str.strip(),
                errors='coerce'
            ) if 'point_id' in col_map else None,
            'z': pd.to_numeric(df[col_map['z']], errors='coerce') if 'z' in col_map else 0,
        }).reset_index(drop=True)
    else:
        # Positional: point_id, y, x, z
        if len(raw.columns) < 3:
            raise ValueError(f"CSV without header must have at least 3 columns (point_id, y, x)")

        raw[0] = raw[0].astype(str).str.replace(r'^\s*CK\s*', '', regex=True).str.strip()

        out = pd.DataFrame({
            'point_id': pd.to_numeric(raw[0], errors='coerce'),
            'y': pd.to_numeric(raw[1], errors='coerce'),
            'x': pd.to_numeric(raw[2], errors='coerce'),
            'z': pd.to_numeric(raw[3], errors='coerce') if len(raw.columns) >= 4 else 0,
        }).reset_index(drop=True)

    if out['point_id'].isnull().all() or 'point_id' not in out.columns or out['point_id'] is None:
        out['point_id'] = out.index + 1

    return out


def compute_deviations(csv_bytes: bytes, geojson_path: str, tif_path: str) -> dict:
    """
    Compute QC point deviations against detection polygons.

    Loads QC points (in TIF's native CRS), finds containing detection polygons,
    computes distance from each point to the polygon centroid (converted to cm),
    and reprojects results to WGS84 for map display.

    Args:
        csv_bytes: Raw CSV file bytes for QC points
        geojson_path: Path to detections GeoJSON (native CRS)
        tif_path: Path to source GeoTIFF (for CRS and bounds)

    Returns:
        Dict with qc_points list and summary stats
    """
    THRESHOLD_CM = 3.0
    FEET_TO_CM = 30.48

    # Load QC points
    qc_df = load_qc_points(csv_bytes)

    # Get TIF CRS and bounds
    with rasterio.open(tif_path) as src:
        tif_crs = src.crs
        tif_bounds = src.bounds

    # Load detection polygons (in native CRS)
    polygons_gdf = gpd.read_file(geojson_path)
    if polygons_gdf.crs is None:
        polygons_gdf = polygons_gdf.set_crs(tif_crs)
    elif str(polygons_gdf.crs) != str(tif_crs):
        polygons_gdf = polygons_gdf.to_crs(tif_crs)

    # Drop rows with missing coordinates
    qc_df = qc_df.dropna(subset=['x', 'y']).reset_index(drop=True)

    if len(qc_df) == 0:
        # Re-read raw CSV to show what was parsed
        raw_df = pd.read_csv(io.BytesIO(csv_bytes), header=None, dtype=str, nrows=3)
        preview = raw_df.to_string(index=False)
        raise ValueError(
            f"No valid QC points after parsing. Check column order (expected: point_id, y, x, z). "
            f"First rows of CSV:\n{preview}"
        )

    # Convert QC points to GeoDataFrame in TIF CRS
    points_gdf = gpd.GeoDataFrame(
        qc_df,
        geometry=gpd.points_from_xy(qc_df['x'], qc_df['y']),
        crs=tif_crs
    )

    # Filter to TIF bounds
    left, bottom, right, top = tif_bounds
    in_bounds = (
        (points_gdf.geometry.x >= left) &
        (points_gdf.geometry.x <= right) &
        (points_gdf.geometry.y >= bottom) &
        (points_gdf.geometry.y <= top)
    )
    points_filtered = points_gdf[in_bounds].copy()

    # Compute deviations
    results = []
    for _, point in points_filtered.iterrows():
        containing = polygons_gdf[polygons_gdf.contains(point.geometry)]

        if not containing.empty:
            # Find the closest centroid among containing polygons
            best_dist = float('inf')
            best_poly_id = None
            best_centroid = None

            for poly_idx, poly in containing.iterrows():
                centroid = poly.geometry.centroid
                dist = point.geometry.distance(centroid)
                if dist < best_dist:
                    best_dist = dist
                    best_poly_id = poly_idx
                    best_centroid = centroid

            deviation_cm = best_dist * FEET_TO_CM

            results.append({
                'point_id': int(point['point_id']) if pd.notna(point['point_id']) else 0,
                'matched': True,
                'deviation_cm': round(deviation_cm, 2),
                'exceeds_threshold': deviation_cm > THRESHOLD_CM,
                'point_geom': point.geometry,
                'centroid_geom': best_centroid,
            })
        else:
            results.append({
                'point_id': int(point['point_id']) if pd.notna(point['point_id']) else 0,
                'matched': False,
                'deviation_cm': None,
                'exceeds_threshold': False,
                'point_geom': point.geometry,
                'centroid_geom': None,
            })

    # Batch reproject QC points to WGS84
    pts_gdf = gpd.GeoDataFrame(
        [{'idx': i, 'geometry': r['point_geom']} for i, r in enumerate(results)],
        crs=tif_crs
    ).to_crs('EPSG:4326')

    # Batch reproject matched centroids to WGS84
    centroid_rows = [
        {'idx': i, 'geometry': r['centroid_geom']}
        for i, r in enumerate(results) if r['centroid_geom'] is not None
    ]
    if centroid_rows:
        ct_gdf = gpd.GeoDataFrame(centroid_rows, crs=tif_crs).to_crs('EPSG:4326')
        centroid_map = {int(row['idx']): row.geometry for _, row in ct_gdf.iterrows()}
    else:
        centroid_map = {}

    # Build output
    qc_points_out = []
    for i, r in enumerate(results):
        pt_wgs = pts_gdf.loc[pts_gdf['idx'] == i].geometry.iloc[0]
        entry = {
            'point_id': r['point_id'],
            'lat': round(pt_wgs.y, 8),
            'lng': round(pt_wgs.x, 8),
            'matched': r['matched'],
            'deviation_cm': r['deviation_cm'],
            'exceeds_threshold': r['exceeds_threshold'],
            'centroid_lat': None,
            'centroid_lng': None,
        }

        if i in centroid_map:
            ct_wgs = centroid_map[i]
            entry['centroid_lat'] = round(ct_wgs.y, 8)
            entry['centroid_lng'] = round(ct_wgs.x, 8)

        qc_points_out.append(entry)

    # Summary stats
    matched = [r for r in results if r['matched']]
    deviations = [r['deviation_cm'] for r in matched if r['deviation_cm'] is not None]

    summary = {
        'total_qc_points': len(points_filtered),
        'matched_points': len(matched),
        'unmatched_points': len(points_filtered) - len(matched),
        'avg_deviation_cm': round(sum(deviations) / len(deviations), 2) if deviations else 0,
        'max_deviation_cm': round(max(deviations), 2) if deviations else 0,
        'count_exceeding_3cm': sum(1 for d in deviations if d > THRESHOLD_CM),
        'threshold_cm': THRESHOLD_CM,
    }

    return {
        'qc_points': qc_points_out,
        'summary': summary,
    }


def generate_deviation_report(csv_bytes: bytes, geojson_path: str, tif_path: str,
                              log_callback=None) -> bytes:
    """
    Generate an Excel deviation report with embedded visualization.

    Sheet 1 "Deviation Report": table with Point ID, Y, X, Z (dummy), Deviation, Status
    plus a matplotlib plot of the ortho zoomed to QC bounding box with markers and lines.
    Sheet 2 "Detected Centroids (WGS84)": matched centroids in WGS84 with dummy ELEV.

    Returns:
        Excel workbook as bytes
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.ticker import ScalarFormatter
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import PatternFill, Font, Alignment

    def log(msg):
        if log_callback:
            log_callback(msg)

    THRESHOLD_CM = 3.0
    FEET_TO_CM = 30.48

    # --- Load data (reuse existing helpers) ---
    log("Loading QC points...")
    qc_df = load_qc_points(csv_bytes)

    with rasterio.open(tif_path) as src:
        tif_crs = src.crs
        tif_bounds = src.bounds

    log(f"Loading detection polygons...")
    polygons_gdf = gpd.read_file(geojson_path)
    if polygons_gdf.crs is None:
        polygons_gdf = polygons_gdf.set_crs(tif_crs)
    elif str(polygons_gdf.crs) != str(tif_crs):
        polygons_gdf = polygons_gdf.to_crs(tif_crs)
    log(f"Loaded {len(polygons_gdf)} detection polygons.")

    qc_df = qc_df.dropna(subset=['x', 'y']).reset_index(drop=True)

    points_gdf = gpd.GeoDataFrame(
        qc_df,
        geometry=gpd.points_from_xy(qc_df['x'], qc_df['y']),
        crs=tif_crs
    )

    # Filter to TIF bounds
    left, bottom, right, top = tif_bounds
    in_bounds = (
        (points_gdf.geometry.x >= left) &
        (points_gdf.geometry.x <= right) &
        (points_gdf.geometry.y >= bottom) &
        (points_gdf.geometry.y <= top)
    )
    points_filtered = points_gdf[in_bounds].copy()

    # --- Compute deviations ---
    log(f"Computing deviations for {len(points_filtered)} QC points...")
    results = []
    for _, point in points_filtered.iterrows():
        containing = polygons_gdf[polygons_gdf.contains(point.geometry)]
        if not containing.empty:
            best_dist = float('inf')
            best_centroid = None
            for _, poly in containing.iterrows():
                centroid = poly.geometry.centroid
                dist = point.geometry.distance(centroid)
                if dist < best_dist:
                    best_dist = dist
                    best_centroid = centroid

            deviation_cm = best_dist * FEET_TO_CM
            results.append({
                'point_id': int(point['point_id']) if pd.notna(point['point_id']) else 0,
                'y': point['y'],
                'x': point['x'],
                'deviation_cm': round(deviation_cm, 2),
                'exceeds': deviation_cm > THRESHOLD_CM,
                'centroid_x': best_centroid.x,
                'centroid_y': best_centroid.y,
                'point_geom': point.geometry,
                'centroid_geom': best_centroid,
            })

    # --- Sheet 1: Deviation Report ---
    log(f"Building deviation table ({len(results)} matched points)...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Deviation Report"

    red_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
    header_font = Font(bold=True)
    headers = ["Point ID", "Y", "X", "Deviation (cm)", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r['point_id'])
        ws.cell(row=row_idx, column=2, value=r['y'])
        ws.cell(row=row_idx, column=3, value=r['x'])
        # ws.cell(row=row_idx, column=4, value=round(166 + random.randint(0, 9) / 10, 1))
        ws.cell(row=row_idx, column=4, value=r['deviation_cm'])
        ws.cell(row=row_idx, column=5, value="EXCEED" if r['exceeds'] else "OK")

        if r['exceeds']:
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = red_fill

    # Average deviation row
    if results:
        deviations = [r['deviation_cm'] for r in results]
        avg_row = len(results) + 3
        ws.cell(row=avg_row, column=4, value="Average deviation (cm)")
        ws.cell(row=avg_row, column=4).font = Font(bold=True)
        ws.cell(row=avg_row, column=5, value=round(sum(deviations) / len(deviations), 2))

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10

    # --- Matplotlib visualization (notebook-style, zoomed to QC area) ---
    log("Generating visualization (cropping ortho to QC area)...")
    if results:
        # Compute bounding box of QC points with padding
        all_x = [r['x'] for r in results] + [r['centroid_x'] for r in results]
        all_y = [r['y'] for r in results] + [r['centroid_y'] for r in results]
        bbox_left = min(all_x)
        bbox_right = max(all_x)
        bbox_bottom = min(all_y)
        bbox_top = max(all_y)

        pad_x = (bbox_right - bbox_left) * 0.15 + 2
        pad_y = (bbox_top - bbox_bottom) * 0.15 + 2
        bbox_left -= pad_x
        bbox_right += pad_x
        bbox_bottom -= pad_y
        bbox_top += pad_y

        # Clamp to TIF bounds
        bbox_left = max(bbox_left, tif_bounds.left)
        bbox_right = min(bbox_right, tif_bounds.right)
        bbox_bottom = max(bbox_bottom, tif_bounds.bottom)
        bbox_top = min(bbox_top, tif_bounds.top)

        # Read cropped ortho via rasterio windowed read
        with rasterio.open(tif_path) as src:
            window = from_bounds(bbox_left, bbox_bottom, bbox_right, bbox_top, src.transform)
            data = src.read([1, 2, 3], window=window)
            win_transform = rasterio.windows.transform(window, src.transform)

        # Build extent for imshow
        img_h, img_w = data.shape[1], data.shape[2]
        extent_left = win_transform.c
        extent_top = win_transform.f
        extent_right = extent_left + win_transform.a * img_w
        extent_bottom = extent_top + win_transform.e * img_h

        # Downsample to reasonable size for embedding (max 1500px on longest side)
        img_array = np.moveaxis(data, 0, 2).astype(np.uint8)
        max_dim = 1500
        if max(img_h, img_w) > max_dim:
            scale = max_dim / max(img_h, img_w)
            new_w = int(img_w * scale)
            new_h = int(img_h * scale)
            pil_crop = Image.fromarray(img_array)
            pil_crop = pil_crop.resize((new_w, new_h), Image.Resampling.LANCZOS)
            img_array = np.array(pil_crop)

        qc_ok = [r for r in results if not r['exceeds']]
        qc_exceed = [r for r in results if r['exceeds']]

        fig, ax = plt.subplots(figsize=(8, 6), constrained_layout=True)
        ax.imshow(img_array, extent=[extent_left, extent_right, extent_bottom, extent_top],
                  aspect='equal', origin='upper')

        # QC points within threshold (white scatter)
        if qc_ok:
            ax.scatter([r['x'] for r in qc_ok], [r['y'] for r in qc_ok],
                       c='white', s=20, label='QC Points', edgecolor='black', zorder=3)

        # QC points exceeding threshold (red scatter + labels)
        if qc_exceed:
            ax.scatter([r['x'] for r in qc_exceed], [r['y'] for r in qc_exceed],
                       c='red', s=40, label='Deviations > 3 cm', edgecolor='black', zorder=4)

            for r in qc_exceed:
                # Connecting line
                ax.plot([r['x'], r['centroid_x']], [r['y'], r['centroid_y']],
                        linestyle=':', color='red', linewidth=1.5,
                        alpha=0.8, zorder=2)

                # Distance label at midpoint
                mid_x = (r['x'] + r['centroid_x']) / 2
                mid_y = (r['y'] + r['centroid_y']) / 2
                ax.annotate(f"{r['deviation_cm']:.2f} cm", (mid_x, mid_y),
                            fontsize=7, color='red', fontweight='bold',
                            ha='left', va='bottom',
                            bbox=dict(boxstyle='round,pad=0.15', fc='black', alpha=0.5, ec='none'),
                            zorder=5)

                # Point ID label (black text, offset to upper-left with annotation arrow)
                ax.annotate(str(r['point_id']), (r['x'], r['y']),
                            textcoords="offset points", xytext=(-14, 10),
                            ha='center', fontsize=8, color='black', fontweight='bold',
                            bbox=dict(boxstyle='round,pad=0.2', fc='white', alpha=0.8, ec='gray', lw=0.5),
                            arrowprops=dict(arrowstyle='-', color='gray', lw=0.7),
                            zorder=6)

        ax.set_xlabel("Easting (ft.)")
        ax.set_ylabel("Northing (ft.)")
        ax.legend(loc='upper left', bbox_to_anchor=(1.05, 0.95), borderaxespad=0,
                  fontsize=7, markerscale=0.6, handletextpad=0.4, borderpad=0.6)
        ax.xaxis.set_major_formatter(ScalarFormatter(useOffset=False))
        ax.ticklabel_format(useOffset=False, style='plain', axis='x')

        # Simplify x ticks (first and last only, rotated)
        ticks = ax.get_xticks()
        if len(ticks) >= 2:
            ax.set_xticks([ticks[0], ticks[-1]])
            ax.set_xticklabels([f"{t:.0f}" for t in [ticks[0], ticks[-1]]], rotation=90)

        # Save plot to bytes
        plot_buf = io.BytesIO()
        plt.savefig(plot_buf, format='png', dpi=150, bbox_inches='tight')
        plt.close(fig)
        plot_bytes = plot_buf.getvalue()

        # Get actual plot dimensions for correct aspect ratio
        pil_plot = Image.open(io.BytesIO(plot_bytes))
        plot_w, plot_h = pil_plot.size

        # Embed in Excel at cell E1 (matching notebook), width=500
        log("Embedding visualization in workbook...")
        img = XLImage(io.BytesIO(plot_bytes))
        img.width = 500
        img.height = int(500 * (plot_h / plot_w)) if plot_w > 0 else 400
        ws.add_image(img, "E1")

    # --- Sheet 2: Detected Centroids (WGS84) ---
    log("Building WGS84 centroids sheet...")
    ws2 = wb.create_sheet("Detected Centroids (WGS84)")
    ws2.cell(row=1, column=1, value="Y").font = header_font
    ws2.cell(row=1, column=2, value="X").font = header_font
    ws2.cell(row=1, column=3, value="ELEV").font = header_font

    if results:
        # Only matched centroids — reproject to WGS84
        centroid_geoms = [r['centroid_geom'] for r in results]
        ct_gdf = gpd.GeoDataFrame(
            geometry=centroid_geoms, crs=tif_crs
        ).to_crs('EPSG:4326')

        for row_idx, (_, row) in enumerate(ct_gdf.iterrows(), 2):
            ws2.cell(row=row_idx, column=1, value=round(row.geometry.y, 8))
            ws2.cell(row=row_idx, column=2, value=round(row.geometry.x, 8))
            ws2.cell(row=row_idx, column=3, value=round(166 + random.randint(0, 9) / 10, 1))

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 10

    # --- Sheet 3: Deviation Snapshots (only for exceeding points) ---
    exceed_results = [r for r in results if r['exceeds']]
    if exceed_results:
        log(f"Generating {len(exceed_results)} deviation snapshots...")
        ws3 = wb.create_sheet("Deviation Snapshots")
        ws3.cell(row=1, column=1, value="Point ID").font = header_font
        ws3.cell(row=1, column=2, value="Deviation (cm)").font = header_font
        ws3.cell(row=1, column=3, value="Snapshot").font = header_font
        ws3.column_dimensions['A'].width = 12
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['C'].width = 18

        current_row = 2
        for r in exceed_results:
            # Crop ortho centered on midpoint of QC point and centroid
            mid_x = (r['x'] + r['centroid_x']) / 2
            mid_y = (r['y'] + r['centroid_y']) / 2

            # Adaptive padding: scale to the distance between points so they're clearly visible
            dist_ft = r['deviation_cm'] / FEET_TO_CM  # deviation in feet
            pad = max(dist_ft * 5, 0.5)  # at least 0.5 ft, otherwise 5x the deviation distance

            crop_left = max(mid_x - pad, tif_bounds.left)
            crop_right = min(mid_x + pad, tif_bounds.right)
            crop_bottom = max(mid_y - pad, tif_bounds.bottom)
            crop_top = min(mid_y + pad, tif_bounds.top)

            with rasterio.open(tif_path) as src:
                win = from_bounds(crop_left, crop_bottom, crop_right, crop_top, src.transform)
                snap_data = src.read([1, 2, 3], window=win)
                snap_transform = rasterio.windows.transform(win, src.transform)

            sh, sw = snap_data.shape[1], snap_data.shape[2]
            snap_arr = np.moveaxis(snap_data, 0, 2).astype(np.uint8)

            # Downsample to ~300px for sharper snapshots
            if max(sh, sw) > 300:
                sc = 300 / max(sh, sw)
                pil_snap = Image.fromarray(snap_arr).resize(
                    (int(sw * sc), int(sh * sc)), Image.Resampling.LANCZOS)
                snap_arr = np.array(pil_snap)

            # Build extent
            s_left = snap_transform.c
            s_top = snap_transform.f
            s_right = s_left + snap_transform.a * sw
            s_bottom = s_top + snap_transform.e * sh

            # Plot snapshot
            fig_s, ax_s = plt.subplots(figsize=(2.5, 2.5))
            ax_s.imshow(snap_arr, extent=[s_left, s_right, s_bottom, s_top],
                        aspect='equal', origin='upper')

            # Red QC point
            ax_s.scatter(r['x'], r['y'], c='red', s=80, edgecolors='black',
                         linewidths=0.8, zorder=4, label='QC Point')
            # White centroid
            ax_s.scatter(r['centroid_x'], r['centroid_y'], c='white', s=60,
                         edgecolors='black', linewidths=0.8, zorder=3, label='Centroid')
            # Connecting line
            ax_s.plot([r['x'], r['centroid_x']], [r['y'], r['centroid_y']],
                      ':', color='red', linewidth=2, zorder=2)
            # Measurement label (offset well above midpoint so it doesn't occlude markers)
            ax_s.annotate(f"{r['deviation_cm']:.2f} cm",
                          (mid_x, mid_y), fontsize=9, color='red', fontweight='bold',
                          ha='center', va='bottom',
                          textcoords="offset points", xytext=(0, 22),
                          bbox=dict(boxstyle='round,pad=0.2', fc='black', alpha=0.7, ec='none'),
                          arrowprops=dict(arrowstyle='-', color='red', lw=0.8, alpha=0.6),
                          zorder=5)

            ax_s.set_xticks([])
            ax_s.set_yticks([])
            ax_s.set_frame_on(False)
            plt.subplots_adjust(left=0, right=1, top=1, bottom=0)

            snap_buf = io.BytesIO()
            plt.savefig(snap_buf, format='png', dpi=120, bbox_inches='tight', pad_inches=0.02)
            plt.close(fig_s)
            snap_bytes = snap_buf.getvalue()

            # Write label cells
            ws3.cell(row=current_row, column=1, value=f"{r['point_id']}").font = header_font
            ws3.cell(row=current_row, column=2, value=f"{r['deviation_cm']}")

            # Embed snapshot
            snap_img = XLImage(io.BytesIO(snap_bytes))
            snap_img.width = 150
            snap_img.height = 150
            ws3.add_image(snap_img, f"C{current_row}")

            current_row += 10  # space for image height

    # Save to bytes
    log("Saving workbook...")
    output = io.BytesIO()
    wb.save(output)
    log("Report complete.")
    return output.getvalue()


def cleanup_old_files(directory: str, max_age_hours: int = 1):
    """
    Clean up old uploaded files.

    Args:
        directory: Directory to clean
        max_age_hours: Maximum age in hours
    """
    import time

    if not os.path.exists(directory):
        return

    current_time = time.time()
    max_age_seconds = max_age_hours * 3600

    for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)
        if os.path.isfile(filepath):
            file_age = current_time - os.path.getmtime(filepath)
            if file_age > max_age_seconds:
                try:
                    os.remove(filepath)
                    print(f"Cleaned up old file: {filename}")
                except Exception as e:
                    print(f"Error removing {filename}: {e}")
